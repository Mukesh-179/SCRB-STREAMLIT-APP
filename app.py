"""
app.py — SCRB Crime Intelligence & Analytics Platform (Streamlit prototype)
Run with: streamlit run app.py
"""
import streamlit as st
from io import BytesIO
import re
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import networkx as nx
import backend as b
from datetime import datetime
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from copy import copy
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference

st.set_page_config(page_title="SCRB Crime Intelligence Platform", page_icon="🕵️",
                    layout="wide", initial_sidebar_state="expanded")

# ---------------------------------------------------------------- THEME / CSS
PLOTLY_TEMPLATE = "plotly_dark"
AMBER, TEAL, RED, BLUE, VIOLET, MUTED = "#ffb84d", "#72e0b8", "#ff5d6e", "#6f7dff", "#a855f7", "#9fb3d0"
PALETTE = [VIOLET, RED, BLUE, TEAL, AMBER, "#c084fc", "#f472b6", "#93c5fd", "#a78bfa", "#86efac"]

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');
html, body, [class*="css"]  { font-family: 'Inter', sans-serif; }
.main { background-color: #07060d; }
[data-testid="stAppViewContainer"] { background:
    radial-gradient(ellipse 850px 480px at 85% -10%, rgba(168,85,247,0.24), transparent),
    radial-gradient(ellipse 850px 480px at 15% 0%, rgba(244,114,182,0.12), transparent),
    radial-gradient(ellipse 700px 380px at 50% 100%, rgba(125,211,252,0.08), transparent),
    #07060d; }
[data-testid="stSidebar"] { background: linear-gradient(180deg, #171126 0%, #0b0913 100%); border-right: 1px solid rgba(168,85,247,0.25); }
[data-testid="stSidebarNav"], [data-testid="stSidebarNav"] ul { padding-top: 0.2rem; }
[data-testid="stSidebar"] [data-testid="stRadio"] [role="radiogroup"] label {
    position: relative;
    display: flex;
    align-items: center;
    width: 100%;
    min-height: 4.8rem;
    margin-bottom: 0.35rem;
    padding: 0.9rem 0.95rem 0.9rem 2.8rem;
    border-radius: 16px;
    border: 1px solid rgba(255,255,255,0.06);
    background: rgba(255,255,255,0.02);
    transition: all 0.18s ease;
    box-sizing: border-box;
    line-height: 1.2;
}
[data-testid="stSidebar"] [data-testid="stRadio"] [role="radiogroup"] label:hover {
    border-color: rgba(168,85,247,0.4);
    background: rgba(168,85,247,0.08);
}
[data-testid="stSidebar"] [data-testid="stRadio"] [role="radiogroup"] label > div:first-child {
    display: none !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] [role="radiogroup"] label::before {
    content: "";
    position: absolute;
    left: 0.82rem;
    top: 50%;
    width: 0.95rem;
    height: 0.95rem;
    transform: translateY(-50%);
    border-radius: 3px;
    border: 2px solid rgba(159,179,208,0.78);
    background: rgba(7,11,24,0.95);
    box-shadow: inset 0 0 0 2px rgba(7,11,24,0.35);
}
[data-testid="stSidebar"] [data-testid="stRadio"] [role="radiogroup"] label:has(input:checked) {
    border-color: rgba(168,85,247,0.96);
    background: linear-gradient(135deg, rgba(168,85,247,0.96) 0%, rgba(244,114,182,0.92) 100%);
    box-shadow: 0 12px 28px rgba(168,85,247,0.22), inset 0 0 0 1px rgba(255,255,255,0.16);
    color: #140a1f;
}
[data-testid="stSidebar"] [data-testid="stRadio"] [role="radiogroup"] label:has(input:checked)::before {
    border-color: rgba(255,255,255,0.95);
    background: rgba(255,255,255,0.18);
    box-shadow: 0 0 0 3px rgba(255,255,255,0.14);
}
[data-testid="stSidebar"] [data-testid="stRadio"] [role="radiogroup"] label:has(input:checked)::after {
    content: "";
    position: absolute;
    left: 1.08rem;
    top: 50%;
    width: 0.36rem;
    height: 0.36rem;
    transform: translateY(-50%);
    border-radius: 1px;
    background: #04131a;
}
[data-testid="stSidebar"] [data-testid="stRadio"] [role="radiogroup"] label:has(input:checked) * {
    color: #140a1f !important;
}
h1,h2,h3 { font-family: 'Space Grotesk', sans-serif !important; }
.eyebrow { font-family: 'IBM Plex Mono', monospace; font-size: 11px; letter-spacing: 0.13em; color: #c084fc; text-transform: uppercase; margin-bottom: 2px;}
.subtext { color: #b5c3dd; font-size: 13.5px; max-width: 780px; }
.hero-panel { background: radial-gradient(circle at top left, rgba(168,85,247,0.22), transparent 35%), linear-gradient(135deg, rgba(24,18,38,0.98), rgba(12,11,20,0.98)); border: 1px solid rgba(168,85,247,0.3); border-radius: 24px; padding: 22px 24px; margin-bottom: 16px; box-shadow: 0 0 0 1px rgba(168,85,247,0.08), 0 18px 40px rgba(0,0,0,0.28); }
.info-card { background: linear-gradient(145deg, rgba(29,23,41,0.98), rgba(14,12,21,0.98)); border: 1px solid rgba(168,85,247,0.22); border-radius: 16px; padding: 16px 18px; margin-bottom: 12px; box-shadow: inset 0 1px 0 rgba(255,255,255,0.03), 0 10px 24px rgba(0,0,0,0.22); }
.stack-grid { display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; margin: 0 0 16px; }
.stack-card { background: linear-gradient(145deg, rgba(29,23,41,0.98), rgba(13,11,21,0.98)); border: 1px solid rgba(168,85,247,0.18); border-radius: 18px; padding: 16px 16px 14px; box-shadow: 0 10px 24px rgba(0,0,0,0.18); }
.stack-card.primary { border-color: rgba(168,85,247,0.38); background: linear-gradient(145deg, rgba(168,85,247,0.12), rgba(13,11,21,0.98)); }
.stack-card.backend { border-color: rgba(244,114,182,0.28); background: linear-gradient(145deg, rgba(244,114,182,0.08), rgba(13,11,21,0.98)); }
.stack-card.data { border-color: rgba(125,211,252,0.28); background: linear-gradient(145deg, rgba(125,211,252,0.09), rgba(13,11,21,0.98)); }
.stack-title { font-family: 'IBM Plex Mono', monospace; font-size: 10px; letter-spacing: 0.16em; text-transform: uppercase; color: #9fb3d0; margin-bottom: 8px; }
.stack-value { font-family: 'Space Grotesk', sans-serif; font-size: 20px; color: #f5f7fb; margin-bottom: 4px; }
.stack-note { color: #b5c3dd; font-size: 13px; line-height: 1.45; }
.login-card { background: rgba(8,8,14,0.96); border: 1px solid rgba(168,85,247,0.28); border-radius: 22px; padding: 28px; box-shadow: 0 20px 70px rgba(0,0,0,0.45), 0 0 30px rgba(168,85,247,0.12); }
.user-pill { display:inline-block; padding:6px 10px; border-radius:999px; background:rgba(168,85,247,0.12); color:#e9d5ff; border:1px solid rgba(168,85,247,0.25); font-size:12px; font-family:'IBM Plex Mono', monospace; }
.profile-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 8px 12px; margin-top: 10px; }
.profile-item { padding: 8px 10px; border-radius: 10px; background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.05); }
.profile-label { display: block; font-family:'IBM Plex Mono', monospace; font-size: 10px; letter-spacing: 0.08em; text-transform: uppercase; color: #8fa4c3; margin-bottom: 3px; }
.profile-value { font-size: 13px; color: #f5f7fb; font-weight: 600; }
div[data-testid="stMetric"] { background: linear-gradient(145deg, rgba(29,23,41,0.98), rgba(13,11,21,0.98)); border: 1px solid rgba(168,85,247,0.22); border-radius: 18px; padding: 14px 16px 10px; box-shadow: 0 8px 18px rgba(0,0,0,0.18); }
div[data-testid="stMetricLabel"] { font-family: 'IBM Plex Mono', monospace; font-size: 10.5px; letter-spacing:0.05em; text-transform:uppercase; color:#9fb3d0; }
.badge { display:inline-block; padding:3px 10px; border-radius:20px; font-family:'IBM Plex Mono',monospace; font-size:11px; font-weight:500; border:1px solid; }
.badge-critical{color:#ff8ba1;border-color:rgba(255,139,161,0.45);background:rgba(255,139,161,0.12);}
.badge-high{color:#ffd08a;border-color:rgba(255,208,138,0.45);background:rgba(255,208,138,0.12);}
.badge-moderate{color:#c4b5fd;border-color:rgba(196,181,253,0.45);background:rgba(196,181,253,0.12);}
.badge-low{color:#86efac;border-color:rgba(134,239,172,0.35);background:rgba(134,239,172,0.12);}
.mo-tag{display:inline-block;padding:2px 8px;margin:2px 4px 2px 0;border-radius:3px;background:#16111f;border:1px solid rgba(168,85,247,0.22);font-size:11px;color:#d8c9eb;font-family:'IBM Plex Mono',monospace;}
.confidential { font-family:'IBM Plex Mono',monospace; font-size: 10px; color:#7688a6; letter-spacing:0.06em; }
div.stButton > button { background: linear-gradient(90deg, #a855f7 0%, #f472b6 100%); color: white; border: none; border-radius: 999px; padding: 0.55rem 1rem; box-shadow: 0 6px 20px rgba(168,85,247,0.2); }
div.stButton > button:hover { transform: translateY(-1px); box-shadow: 0 8px 24px rgba(168,85,247,0.28); }
hr { border-color: rgba(168,85,247,0.22) !important; }

.dashboard-shell { display: grid; grid-template-columns: 88px minmax(0, 1fr) 300px; gap: 16px; align-items: start; margin: 18px 0 20px; }
.rail-card, .panel-card, .activity-card { background: linear-gradient(180deg, rgba(28,22,38,0.98), rgba(14,12,22,0.98)); border: 1px solid rgba(168,85,247,0.18); border-radius: 24px; box-shadow: 0 20px 48px rgba(0,0,0,0.25); }
.rail-card { padding: 18px 10px; display: flex; flex-direction: column; gap: 18px; align-items: center; position: sticky; top: 18px; }
.rail-icon { width: 42px; height: 42px; border-radius: 999px; display: grid; place-items: center; color: #f5f3ff; background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.08); font-size: 18px; }
.rail-icon.active { background: linear-gradient(180deg, rgba(168,85,247,0.96), rgba(244,114,182,0.82)); border-color: rgba(255,255,255,0.12); color: #140a1f; box-shadow: 0 14px 28px rgba(168,85,247,0.25); }
.dashboard-main { min-width: 0; }
.dashboard-top-grid { display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 14px; margin-bottom: 16px; }
.hero-stat-card { background: linear-gradient(180deg, rgba(30,24,42,0.98), rgba(15,12,23,0.98)); border: 1px solid rgba(168,85,247,0.18); border-radius: 24px; padding: 18px 18px 16px; min-height: 140px; box-shadow: 0 16px 40px rgba(0,0,0,0.22); display: flex; flex-direction: column; justify-content: space-between; }
.hero-stat-number { font-family: 'Space Grotesk', sans-serif; font-size: 38px; line-height: 1; color: #ffffff; letter-spacing: -0.04em; }
.hero-stat-title { font-family: 'Space Grotesk', sans-serif; font-size: 18px; color: #f5f3ff; margin-top: 10px; }
.hero-stat-sub { color: #b7aacd; font-size: 13px; margin-top: 10px; }
.hero-stat-pill { display: inline-flex; align-items: center; gap: 6px; align-self: flex-start; border-radius: 999px; padding: 6px 10px; font-family: 'IBM Plex Mono', monospace; font-size: 11px; letter-spacing: 0.04em; }
.hero-stat-pill.good { background: rgba(134,239,172,0.12); color: #86efac; }
.hero-stat-pill.warn { background: rgba(255,139,161,0.12); color: #ff8ba1; }
.panel-card { padding: 18px; }
.panel-head { display: flex; justify-content: space-between; gap: 12px; align-items: start; margin-bottom: 14px; }
.panel-title { font-family: 'Space Grotesk', sans-serif; font-size: 26px; color: #f7f2ff; margin: 0; }
.panel-subtitle { color: #b8abcf; font-size: 13.5px; margin-top: 6px; max-width: 520px; }
.panel-toolbar { display: flex; align-items: center; gap: 10px; color: #d8c9eb; font-size: 13px; }
.panel-chip { padding: 7px 12px; border-radius: 999px; background: rgba(255,255,255,0.03); border: 1px solid rgba(168,85,247,0.18); color: #e9d5ff; }
.activity-card { padding: 18px; }
.activity-list { margin-top: 14px; display: flex; flex-direction: column; gap: 10px; }
.activity-item { display: flex; justify-content: space-between; align-items: center; gap: 12px; padding: 10px 12px; border-radius: 16px; background: rgba(0,0,0,0.3); border: 1px solid rgba(168,85,247,0.12); }
.activity-label { color: #f4efff; font-size: 13px; }
.activity-check { width: 22px; height: 22px; border-radius: 999px; display: grid; place-items: center; border: 1px solid rgba(255,255,255,0.14); color: #f4efff; font-size: 12px; }
.activity-check.on { background: linear-gradient(180deg, rgba(168,85,247,0.96), rgba(244,114,182,0.82)); border-color: transparent; color: #140a1f; }
.activity-check.off { background: transparent; color: #9ca3af; }
.promo-card { margin-top: 16px; padding: 18px; border-radius: 24px; background: linear-gradient(135deg, rgba(168,85,247,0.88), rgba(96,165,250,0.58)); color: #f8f5ff; position: relative; overflow: hidden; }
.promo-card::after { content: ""; position: absolute; right: -18px; top: -18px; width: 120px; height: 120px; border-radius: 999px; background: rgba(255,255,255,0.12); }
.promo-title { font-family: 'Space Grotesk', sans-serif; font-size: 18px; margin-bottom: 6px; }
.promo-copy { color: rgba(248,245,255,0.82); font-size: 13px; line-height: 1.5; }
.promo-button { margin-top: 18px; background: rgba(255,255,255,0.96); color: #15111f; border-radius: 999px; text-align: center; padding: 12px 16px; font-weight: 600; }
.rail-footer { margin-top: auto; display: flex; flex-direction: column; gap: 12px; width: 100%; }
.rail-footer .rail-icon { width: 46px; height: 46px; }

@media (max-width: 1200px) {
    .dashboard-shell { grid-template-columns: 72px minmax(0, 1fr); }
    .dashboard-side { grid-column: 1 / -1; }
}

@media (max-width: 900px) {
    .dashboard-shell { grid-template-columns: 1fr; }
    .rail-card { position: static; flex-direction: row; flex-wrap: wrap; justify-content: center; }
    .dashboard-top-grid { grid-template-columns: 1fr; }
}
</style>
""", unsafe_allow_html=True)


def px_style(fig, height=380):
    fig.update_layout(template=PLOTLY_TEMPLATE, paper_bgcolor="#171126", plot_bgcolor="#171126",
                       font=dict(family="Inter, sans-serif", color="#ddd6fe", size=12),
                       height=height, margin=dict(l=10, r=10, t=40, b=10),
                       legend=dict(bgcolor="rgba(0,0,0,0)"))
    fig.update_xaxes(gridcolor="rgba(255,255,255,0.06)")
    fig.update_yaxes(gridcolor="rgba(255,255,255,0.06)")
    return fig


def risk_badge(band):
    cls = {"Critical": "badge-critical", "High": "badge-high", "Moderate": "badge-moderate", "Low": "badge-low"}.get(band, "")
    return f'<span class="badge {cls}">{band}</span>'


def normalize_series(series):
    return series.fillna("").astype(str).str.strip().str.replace(r"\.0$", "", regex=True)


def build_kpi_snapshot(df):
    monthly = b.monthly_totals(df).copy()
    monthly["month_dt"] = pd.to_datetime(monthly["month"], format="%b-%y")
    monthly_totals = pd.to_numeric(monthly["total"], errors="coerce").fillna(0)
    peak_index = int(monthly_totals.idxmax())
    highest_month_label = str(monthly.iloc[peak_index]["month"])
    highest_month_total = int(float(monthly_totals.max()))
    avg_monthly = float(monthly["total"].mean())
    total_complaints = int(len(df))
    total_crimes = float(df["Commits"].sum())
    crime_categories = int(df["Major Crime Head"].nunique())

    cats = b.top_categories(df, min(10, crime_categories)).index.tolist()
    labels, series = b.category_monthly(df, cats)
    growth_rows = []
    for cat in cats:
        values = np.asarray(series[cat], dtype=float)
        if len(values) < 6:
            continue
        recent = values[-3:].sum()
        prior = values[-6:-3].sum()
        if prior <= 0:
            continue
        growth_rows.append({"Crime Category": cat, "Recent 3-Month Total": round(recent, 1), "Previous 3-Month Total": round(prior, 1), "Growth %": round(((recent - prior) / prior) * 100, 1)})
    growth_df = pd.DataFrame(growth_rows).sort_values("Growth %", ascending=False) if growth_rows else pd.DataFrame(columns=["Crime Category", "Recent 3-Month Total", "Previous 3-Month Total", "Growth %"])
    fastest = growth_df.iloc[0] if len(growth_df) else pd.Series({"Crime Category": "N/A", "Growth %": 0.0})

    top_categories_df = b.top_categories(df, 8).reset_index()
    top_categories_df.columns = ["Crime Category", "Total Crimes"]
    monthly_chart = monthly[["month", "total"]].copy()

    insights = [
        f"Highest crime month: {highest_month_label} ({highest_month_total} crimes)",
        f"Fastest growing crime: {fastest['Crime Category']} ({fastest['Growth %']:+.1f}%)",
        f"Average monthly crimes: {avg_monthly:.1f}",
        f"Crime categories tracked: {crime_categories}",
    ]

    return {
        "monthly": monthly_chart,
        "highest_month": {"month": highest_month_label, "total": highest_month_total},
        "avg_monthly": avg_monthly,
        "total_complaints": total_complaints,
        "total_crimes": total_crimes,
        "crime_categories": crime_categories,
        "growth_table": growth_df,
        "fastest_growing": fastest,
        "top_categories": top_categories_df,
        "insights": insights,
    }


def build_quality_audit(df):
    raw = df.copy()
    complaint = normalize_series(raw["Complaint Number"])
    major = normalize_series(raw["Major Crime Head"])
    section = normalize_series(raw["Crime Head and Section"])
    minor = normalize_series(raw["Minor Crime Head"])
    month = normalize_series(raw["Month"])
    commits = pd.to_numeric(raw["Commits"], errors="coerce")
    month_valid = pd.to_datetime(month, format="%b-%y", errors="coerce").notna()
    section_valid = section.str.match(r"^Sec\.", na=False)

    missing_by_column = pd.DataFrame({
        "Column": raw.columns,
        "Missing Values": [int(raw[col].isna().sum() + normalize_series(raw[col]).eq("").sum()) for col in raw.columns],
    }).sort_values("Missing Values", ascending=False)

    duplicate_complaint_mask = complaint.ne("") & complaint.duplicated(keep=False)
    duplicate_complaints = raw.loc[duplicate_complaint_mask].copy()
    duplicate_complaints["Complaint Number"] = complaint[duplicate_complaint_mask]
    duplicate_counts = complaint[duplicate_complaint_mask].value_counts().reset_index()
    duplicate_counts.columns = ["Complaint Number", "Occurrences"]

    repeated_records_mask = raw.fillna("").astype(str).duplicated(keep=False)
    repeated_records = raw.loc[repeated_records_mask].copy()

    invalid_mask = (
        complaint.eq("")
        | major.eq("")
        | section.eq("")
        | ~month_valid
        | commits.isna()
        | (commits < 0)
        | ~section_valid
    )
    invalid_rows = raw.loc[invalid_mask].copy()
    invalid_rows["Issue"] = [
        ", ".join(filter(None, [
            "Missing complaint number" if complaint.loc[i] == "" else "",
            "Missing major crime head" if major.loc[i] == "" else "",
            "Missing or malformed section" if (section.loc[i] == "" or not section_valid.loc[i]) else "",
            "Invalid month" if not month_valid.loc[i] else "",
            "Invalid commits" if pd.isna(commits.loc[i]) or commits.loc[i] < 0 else "",
        ]))
        for i in invalid_rows.index
    ]

    issue_summary = pd.DataFrame([
        {"Check": "Duplicate complaint numbers", "Count": int(duplicate_complaints["Complaint Number"].nunique()), "Rows": int(len(duplicate_complaints))},
        {"Check": "Repeated records", "Count": int(repeated_records.shape[0]), "Rows": int(repeated_records.shape[0])},
        {"Check": "Missing values", "Count": int(raw.isna().sum().sum() + sum(normalize_series(raw[col]).eq("").sum() for col in raw.columns)), "Rows": int(len(raw))},
        {"Check": "Invalid months", "Count": int((~month_valid).sum()), "Rows": int((~month_valid).sum())},
        {"Check": "Empty crime categories", "Count": int(major.eq("").sum()), "Rows": int(major.eq("").sum())},
        {"Check": "Section inconsistencies", "Count": int((section.eq("") | ~section_valid).sum()), "Rows": int((section.eq("") | ~section_valid).sum())},
        {"Check": "Invalid data rows", "Count": int(invalid_mask.sum()), "Rows": int(invalid_mask.sum())},
    ])

    return {
        "missing_by_column": missing_by_column,
        "duplicate_counts": duplicate_counts,
        "duplicate_complaints": duplicate_complaints,
        "repeated_records": repeated_records,
        "invalid_rows": invalid_rows,
        "issue_summary": issue_summary,
        "duplicate_complaint_rows": int(duplicate_complaints.shape[0]),
        "repeated_record_rows": int(repeated_records.shape[0]),
        "missing_value_total": int(missing_by_column["Missing Values"].sum()),
        "invalid_month_rows": int((~month_valid).sum()),
        "empty_category_rows": int(major.eq("").sum()),
        "section_inconsistency_rows": int((section.eq("") | ~section_valid).sum()),
        "invalid_data_rows": int(invalid_mask.sum()),
    }


def build_report_bundle(df, meta):
    kpi = build_kpi_snapshot(df)
    quality = build_quality_audit(df)
    summary_rows = pd.DataFrame([
        {"Section": "Overview", "Metric": "Total Complaints", "Value": kpi["total_complaints"], "Detail": "Number of complaint rows"},
        {"Section": "Overview", "Metric": "Total Crimes", "Value": round(kpi["total_crimes"], 1), "Detail": "Sum of commits"},
        {"Section": "KPI", "Metric": "Highest Crime Month", "Value": kpi["highest_month"]["month"], "Detail": f"{int(kpi['highest_month']['total'])} crimes"},
        {"Section": "KPI", "Metric": "Fastest Growing Crime", "Value": kpi["fastest_growing"]["Crime Category"], "Detail": f"{kpi['fastest_growing']['Growth %']:+.1f}%"},
        {"Section": "KPI", "Metric": "Average Monthly Crimes", "Value": round(kpi["avg_monthly"], 1), "Detail": "Monthly mean of total crimes"},
        {"Section": "KPI", "Metric": "Crime Categories", "Value": kpi["crime_categories"], "Detail": "Unique major crime heads"},
        {"Section": "Quality", "Metric": "Duplicate Complaint Numbers", "Value": quality["duplicate_complaint_rows"], "Detail": "Rows with repeated complaint numbers"},
        {"Section": "Quality", "Metric": "Repeated Records", "Value": quality["repeated_record_rows"], "Detail": "Perfect duplicate rows"},
        {"Section": "Quality", "Metric": "Missing Values", "Value": quality["missing_value_total"], "Detail": "Blank or null cells"},
        {"Section": "Quality", "Metric": "Invalid Months", "Value": quality["invalid_month_rows"], "Detail": "Rows with bad month values"},
        {"Section": "Quality", "Metric": "Section Inconsistencies", "Value": quality["section_inconsistency_rows"], "Detail": "Blank or malformed section values"},
        {"Section": "Quality", "Metric": "Invalid Data Rows", "Value": quality["invalid_data_rows"], "Detail": "Rows failing validation rules"},
    ])

    return {"kpi": kpi, "quality": quality, "summary_rows": summary_rows, "meta": meta}


def _plot_drawing(title, labels, values, color):
    drawing = Drawing(500, 220)
    chart = VerticalBarChart()
    chart.x = 35
    chart.y = 30
    chart.width = 430
    chart.height = 150
    chart.data = [list(values)]
    chart.categoryAxis.categoryNames = [str(label) for label in labels]
    chart.categoryAxis.labels.fontSize = 7
    chart.categoryAxis.labels.angle = 35 if len(labels) > 5 else 0
    chart.categoryAxis.labels.dy = -10
    chart.categoryAxis.labels.dx = 5
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = max(values) * 1.2 if len(values) and max(values) > 0 else 1
    chart.bars[0].fillColor = rl_colors.HexColor(color)
    chart.bars[0].strokeColor = rl_colors.HexColor(color)
    drawing.add(chart)
    return drawing


def build_pdf_report(bundle):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=28, bottomMargin=28)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22, textColor=rl_colors.HexColor("#111b2e")))
    styles.add(ParagraphStyle(name="ReportSub", parent=styles["BodyText"], fontName="Helvetica", fontSize=10, leading=13, textColor=rl_colors.HexColor("#334155")))
    story = [
        Paragraph("SCRB Crime Intelligence Report", styles["ReportTitle"]),
        Paragraph(f"Coverage: {bundle['meta']['month_range']} | Districts: {bundle['meta']['n_districts']} | Categories: {bundle['meta']['n_categories']}", styles["ReportSub"]),
        Spacer(1, 10),
    ]

    summary_table = [["Section", "Metric", "Value", "Detail"]] + bundle["summary_rows"].values.tolist()
    summary = Table(summary_table, repeatRows=1, colWidths=[68, 135, 90, 180])
    summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#111b2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.whitesmoke, rl_colors.HexColor("#eef2ff")]),
    ]))
    story.extend([summary, Spacer(1, 12), Paragraph("Key insights", styles["Heading2"])])
    for insight in bundle["kpi"]["insights"]:
        story.append(Paragraph(f"• {insight}", styles["ReportSub"]))

    story.extend([Spacer(1, 10), Paragraph("Monthly crime trend", styles["Heading2"]), _plot_drawing("Monthly trend", bundle["kpi"]["monthly"]["month"].tolist()[-8:], bundle["kpi"]["monthly"]["total"].tolist()[-8:], TEAL)])
    story.extend([Spacer(1, 10), Paragraph("Top crime categories", styles["Heading2"]), _plot_drawing("Top categories", bundle["kpi"]["top_categories"]["Crime Category"].tolist()[:6], bundle["kpi"]["top_categories"]["Total Crimes"].tolist()[:6], AMBER)])

    doc.build(story)
    return buffer.getvalue()


def build_excel_report(bundle):
    wb = Workbook()
    ws = wb.create_sheet("Summary", 0)
    ws.append(["Section", "Metric", "Value", "Detail"])
    for row in bundle["summary_rows"].itertuples(index=False):
        ws.append(list(row))
    for cell in ws[1]:
            header_font = copy(cell.font)
            header_font.bold = True
            cell.font = header_font

    monthly_ws = wb.create_sheet("Monthly Crimes")
    monthly_ws.append(["Month", "Total Crimes"])
    for row in bundle["kpi"]["monthly"].itertuples(index=False):
        monthly_ws.append([row.month, float(row.total)])
    monthly_chart = BarChart()
    monthly_chart.title = "Monthly Crime Trend"
    monthly_chart.y_axis.title = "Crimes"
    monthly_chart.x_axis.title = "Month"
    monthly_chart.height = 7
    monthly_chart.width = 13
    data = Reference(monthly_ws, min_col=2, min_row=1, max_row=monthly_ws.max_row)
    cats = Reference(monthly_ws, min_col=1, min_row=2, max_row=monthly_ws.max_row)
    monthly_chart.add_data(data, titles_from_data=True)
    monthly_chart.set_categories(cats)
    monthly_ws.add_chart(monthly_chart, "D2")

    cat_ws = wb.create_sheet("Top Categories")
    cat_ws.append(["Crime Category", "Total Crimes"])
    for row in bundle["kpi"]["top_categories"].itertuples(index=False):
        cat_ws.append([row[0], float(row[1])])
    cat_chart = BarChart()
    cat_chart.title = "Top Crime Categories"
    cat_chart.y_axis.title = "Crimes"
    cat_chart.x_axis.title = "Category"
    cat_chart.height = 7
    cat_chart.width = 13
    data = Reference(cat_ws, min_col=2, min_row=1, max_row=cat_ws.max_row)
    cats = Reference(cat_ws, min_col=1, min_row=2, max_row=cat_ws.max_row)
    cat_chart.add_data(data, titles_from_data=True)
    cat_chart.set_categories(cats)
    cat_ws.add_chart(cat_chart, "D2")

    quality_ws = wb.create_sheet("Quality Audit")
    quality_ws.append(["Check", "Count", "Rows"])
    for row in bundle["quality"]["issue_summary"].itertuples(index=False):
        quality_ws.append(list(row))
    quality_chart = BarChart()
    quality_chart.title = "Data Quality Issues"
    quality_chart.height = 7
    quality_chart.width = 13
    data = Reference(quality_ws, min_col=2, min_row=1, max_row=quality_ws.max_row)
    cats = Reference(quality_ws, min_col=1, min_row=2, max_row=quality_ws.max_row)
    quality_chart.add_data(data, titles_from_data=True)
    quality_chart.set_categories(cats)
    quality_ws.add_chart(quality_chart, "E2")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_csv_summary(bundle):
    return bundle["summary_rows"].to_csv(index=False).encode("utf-8")


def build_quality_report_text(audit):
    lines = [
        "SCRB Data Quality Report",
        "",
        "Issue summary:",
    ]
    for row in audit["issue_summary"].itertuples(index=False):
        lines.append(f"- {row.Check}: {row.Count} ({row.Rows} rows)")
    lines.extend([
        "",
        f"Duplicate complaint rows: {audit['duplicate_complaint_rows']}",
        f"Repeated records: {audit['repeated_record_rows']}",
        f"Missing values: {audit['missing_value_total']}",
        f"Invalid months: {audit['invalid_month_rows']}",
        f"Empty crime categories: {audit['empty_category_rows']}",
        f"Section inconsistencies: {audit['section_inconsistency_rows']}",
        f"Invalid data rows: {audit['invalid_data_rows']}",
    ])
    return "\n".join(lines).encode("utf-8")


def build_issue_ticket(issue_type, priority, summary, details, affected_rows):
    ticket_id = f"ISSUE-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    ticket = pd.DataFrame([
        {
            "ticket_id": ticket_id,
            "issue_type": issue_type,
            "priority": priority,
            "summary": summary,
            "details": details,
            "affected_rows": affected_rows,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    ])
    return ticket_id, ticket


def render_stack_view(meta):
        st.markdown(
                f"""
                <div class="stack-grid">
                    <div class="stack-card primary">
                        <div class="stack-title">1 · Overview + 2 · Taxonomy</div>
                        <div class="stack-value">Command Snapshot</div>
                        <div class="stack-note">Live totals, trend direction, and the crime-head explorer give supervisors a quick read on the current situation.</div>
                    </div>
                    <div class="stack-card backend">
                        <div class="stack-title">3 · Geospatial + 4 · Clustering</div>
                        <div class="stack-value">Hotspot Intelligence</div>
                        <div class="stack-note">Map pressure, hotspot clusters, and district ranking highlight where patrol focus should move next.</div>
                    </div>
                    <div class="stack-card data">
                        <div class="stack-title">5 to 11 · Signals</div>
                        <div class="stack-value">Prediction + Response</div>
                        <div class="stack-note">Anomalies, forecasting, link analysis, socio-economic context, deployment, and the command center round out the full workflow.</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
        )


DEMO_USERS = {
    "ops": {"password": "SCRB2026", "role": "Operations Supervisor", "unit": "State Intelligence Cell", "grade": "Inspector", "level": "Level 3"},
    "admin": {"password": "Admin@SCRB", "role": "Director of SCRB", "unit": "Command Center", "grade": "Superintendent", "level": "Level 5"},
    "analyst": {"password": "Analyst2026", "role": "Crime Analyst", "unit": "Data Fusion Desk", "grade": "Sub-Inspector", "level": "Level 2"},
}


def ensure_authentication():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
        st.session_state.current_user = ""
        st.session_state.user_role = ""
        st.session_state.unit = ""
        st.session_state.police_grade = ""
        st.session_state.police_level = ""

    if st.session_state.authenticated:
        return True

    st.markdown("""
    <div class="hero-panel" style="max-width: 900px; margin: 32px auto 18px;">
      <div class="eyebrow">Secure access</div>
      <h1 style="margin-bottom: 6px;">SCRB Crime Intelligence Command Portal</h1>
      <div class="subtext">Real-time crime monitoring, analyst workflows, and operational response planning in one secure workspace.</div>
    </div>
    """, unsafe_allow_html=True)

    with st.container():
        col1, col2 = st.columns([1.1, 0.9])
        with col1:
            st.markdown("""
            <div class="login-card">
              <div class="eyebrow">Access required</div>
              <h2 style="margin-top: 4px; margin-bottom: 6px;">Sign in to continue</h2>
              <div class="subtext">Use one of the demo accounts below to explore the experience.</div>
              <div class="info-card" style="margin-top: 14px;">
                <div class="confidential">Demo credentials</div>
                <div style="margin-top: 6px;"><b>ops</b> / SCRB2026</div>
                <div><b>admin</b> / Admin@SCRB</div>
                <div><b>analyst</b> / Analyst2026</div>
              </div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown("""
            <div class="login-card">
              <div class="eyebrow">Authentication</div>
              <h3 style="margin-top: 4px;">Operational sign-in</h3>
            </div>
            """, unsafe_allow_html=True)
            with st.form("login_form"):
                username = st.text_input("Username", placeholder="Enter your username")
                password = st.text_input("Password", type="password", placeholder="Enter your password")
                submitted = st.form_submit_button("Enter command center", use_container_width=True)
            if submitted:
                account = DEMO_USERS.get(username.lower())
                if account and account["password"] == password:
                    st.session_state.authenticated = True
                    st.session_state.current_user = username.lower()
                    st.session_state.user_role = account["role"]
                    st.session_state.unit = account["unit"]
                    st.session_state.police_grade = account["grade"]
                    st.session_state.police_level = account["level"]
                    st.success("Access granted. Welcome to the SCRB operations workspace.")
                    st.rerun()
                else:
                    st.error("Invalid credentials. Please use one of the demo accounts shown.")

    st.stop()


ensure_authentication()

# ---------------------------------------------------------------- DATA LOAD
df = b.load_raw()
meta = b.meta_summary()
ddf, district_stations, heat_df, dmc = b.build_district_layer()

# ---------------------------------------------------------------- SIDEBAR
with st.sidebar:
    st.markdown('<div class="eyebrow">Karnataka State Police🚨 &middot; SCRB</div>', unsafe_allow_html=True)
    st.markdown("### Crime Intelligence Platform⚖️")
    st.caption("Secure operations workspace")
    if st.session_state.authenticated:
                st.markdown(f"<div class='user-pill'>👤 {st.session_state.current_user.upper()} · {st.session_state.user_role} · {st.session_state.police_grade}</div>", unsafe_allow_html=True)
                st.markdown(
                        f"""
                        <div class="info-card" style="margin-top: 10px;">
                            <div class="confidential">PERSONNEL DASHBOARD</div>
                            <div class="profile-grid">
                                <div class="profile-item">
                                    <span class="profile-label">Grade</span>
                                    <span class="profile-value">{st.session_state.police_grade or 'N/A'}</span>
                                </div>
                                <div class="profile-item">
                                    <span class="profile-label">Level</span>
                                    <span class="profile-value">{st.session_state.police_level or 'N/A'}</span>
                                </div>
                                <div class="profile-item">
                                    <span class="profile-label">Role</span>
                                    <span class="profile-value">{st.session_state.user_role or 'N/A'}</span>
                                </div>
                                <div class="profile-item">
                                    <span class="profile-label">Unit</span>
                                    <span class="profile-value">{st.session_state.unit or 'N/A'}</span>
                                </div>
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                )
    st.markdown("---")
    section = st.radio("Navigate", [
        "1 · Overview",
        "2 · Taxonomy Explorer",
        "3 · Geospatial Hotspots",
        "4 · Hotspot Clustering (ML)",
        "5 · Trend Anomalies (Statistical)",
        "6 · Anomaly Detection (ML)",
        "7 · Predictive Forecasting",
        "8 · Network & Link Analysis",
        "9 · MO Signature Matching",
        "10· Socio-Economic Overlay",
        "11 · Resource Deployment",
        "12 · Operations Command Center",
        "13 · Duplicate Complaint Detection",
        "14 · Crime Report Generator",
        "15 · KPI Dashboard",
        "16 · Data Quality Dashboard",
        "17 · Dataset Agent Analysis",
    ], label_visibility="collapsed")
    st.markdown(
        """
        <div class="info-card" style="margin-top: 10px;">
          <div class="confidential">APP LAYERS</div>
          <div style="margin-top: 6px; color: #f5f7fb; font-size: 13px;">Frontend: Streamlit UI</div>
          <div style="color: #f5f7fb; font-size: 13px;">Backend: backend.py analytics engine</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("---")
    st.markdown("<div class='info-card'>", unsafe_allow_html=True)
    st.markdown(f"<div class='confidential'>UNIT: {st.session_state.unit or 'N/A'}<br>SYNC: {datetime.now().strftime('%b %d, %H:%M')}<br>RECORDS: {meta['total_records']:,}<br>COVERAGE: {meta['month_range']}<br>CATEGORIES: {meta['n_categories']}<br>DISTRICTS: {meta['n_districts']}</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)
    if st.button("Logout", use_container_width=True):
        st.session_state.authenticated = False
        st.session_state.current_user = ""
        st.session_state.user_role = ""
        st.session_state.unit = ""
        st.session_state.police_grade = ""
        st.session_state.police_level = ""
        st.rerun()

st.markdown(f"""
<div style="display:flex;justify-content:space-between;align-items:flex-end;margin-bottom:6px;">
  <div></div>
  <div style="text-align:right;">
    <span class="confidential">LIVE SYNC · {meta['month_range']} · {meta['n_districts']} DISTRICTS TRACKED</span>
  </div>
</div>
""", unsafe_allow_html=True)

section_num = int(section.split("·", 1)[0].strip().split()[0])

# ============================================================
# 1 · OVERVIEW
# ============================================================
if section_num == 1:
        st.markdown('<div class="eyebrow">Operations Control / 01</div>', unsafe_allow_html=True)
        st.markdown(
                f"""
                <div class="hero-panel">
            <div style="display:flex; justify-content:space-between; align-items:flex-end; gap:16px; flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Crime Detection Dashboard</div>
                <h1 style="margin: 4px 0 6px; font-size: 36px;">Detection Map & Command View</h1>
                <div class="subtext">Monitor recent threats, active cases, and district pressure in a dark command-room layout inspired by the reference design.</div>
            </div>
            <div class="info-card" style="min-width: 260px; margin: 0;">
                <div class="confidential">ACTIVE UNIT</div>
                <div style="font-size: 16px; color: #e9d5ff; margin-top: 4px;">{st.session_state.unit or 'SCRB Operations Desk'}</div>
                <div style="margin-top: 8px; color:#b5c3dd; font-size: 13px;">{meta['month_range']} · {meta['n_districts']} districts · {meta['n_categories']} categories</div>
                <div style="margin-top: 8px; color:#86efac; font-size: 13px; font-family:'IBM Plex Mono', monospace;">{st.session_state.police_grade or 'Police Grade N/A'} · {st.session_state.police_level or 'Level N/A'}</div>
            </div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
        )

        monthly = b.monthly_totals(df)
        recent3 = monthly["total"].tail(3).sum()
        prior3 = monthly["total"].tail(6).head(3).sum()
        delta = (recent3 - prior3) / prior3 * 100 if prior3 else 0
        hotspot_count = int(ddf["is_hotspot"].sum())
        alert_count = len(b.zscore_anomalies(df))
        filed_issues = len(st.session_state.get("filed_issues", []))
        threat_score = alert_count + hotspot_count

        top_categories = b.top_categories(df, 8).reset_index()
        top_categories.columns = ["category", "total"]

        main_col, side_col = st.columns([3.95, 1.25], gap="large")
        with main_col:
            top_left, top_mid, top_right = st.columns(3)
            with top_left:
                st.markdown(
                    f"""
                    <div class="hero-stat-card">
                        <div>
                            <div class="hero-stat-number">{threat_score}</div>
                            <div class="hero-stat-title">Threats detected</div>
                            <div class="hero-stat-sub">{alert_count} anomalies + {hotspot_count} hotspot districts</div>
                        </div>
                        <div class="hero-stat-pill warn">+ {delta:+.1f}% last quarter</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            with top_mid:
                st.markdown(
                    f"""
                    <div class="hero-stat-card">
                        <div>
                            <div class="hero-stat-number">{filed_issues or 4}</div>
                            <div class="hero-stat-title">New cases filed</div>
                            <div class="hero-stat-sub">Agent tickets and data-quality reports opened this session</div>
                        </div>
                        <div class="hero-stat-pill good">Session log active</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            with top_right:
                st.markdown(
                    f"""
                    <div class="hero-stat-card">
                        <div>
                            <div class="hero-stat-number">{hotspot_count}</div>
                            <div class="hero-stat-title">Potential threats</div>
                            <div class="hero-stat-sub">Districts currently above the hotspot threshold</div>
                        </div>
                        <div class="hero-stat-pill warn">Command watch</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            st.markdown(
                """
                <div class="panel-card" style="margin-top: 14px;">
                    <div class="panel-head">
                        <div>
                            <div class="panel-title">Detection Map</div>
                            <div class="panel-subtitle">Monitor and inspect all the crimes or potential threats in the map.</div>
                        </div>
                        <div class="panel-toolbar">
                            <span class="panel-chip">Threats</span>
                            <span class="panel-chip">Secured</span>
                            <span class="panel-chip">Roads</span>
                        </div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            show_area = st.selectbox("Show area", ["By Risk", "By Trend", "By District"], key="overview_show_area")

            plot_df = ddf.sort_values("risk_score", ascending=False).head(18).copy()
            if show_area == "By Trend":
                fig = px.scatter_map(
                    plot_df,
                    lat="lat",
                    lon="lng",
                    size="recent_3mo",
                    color="pct_change_vs_avg",
                    hover_name="district",
                    hover_data={"recent_3mo": ":.1f", "pct_change_vs_avg": ":.1f", "lat": False, "lng": False},
                    color_continuous_scale=[[0, "#60a5fa"], [0.5, "#a855f7"], [1, "#f472b6"]],
                    size_max=30,
                    zoom=6,
                    center={"lat": 15.3, "lon": 75.7},
                )
            elif show_area == "By District":
                plot_df["district_bucket"] = np.where(plot_df["is_hotspot"], "Threat", np.where(plot_df["pct_change_vs_avg"] > 0, "Watch", "Stable"))
                fig = px.scatter_map(
                    plot_df,
                    lat="lat",
                    lon="lng",
                    size="recent_3mo",
                    color="district_bucket",
                    hover_name="district",
                    hover_data={"recent_3mo": ":.1f", "risk_score": ":.1f", "lat": False, "lng": False},
                    color_discrete_map={"Threat": "#f472b6", "Watch": "#a855f7", "Stable": "#86efac"},
                    size_max=30,
                    zoom=6,
                    center={"lat": 15.3, "lon": 75.7},
                )
            else:
                fig = px.scatter_map(
                    plot_df,
                    lat="lat",
                    lon="lng",
                    size="recent_3mo",
                    color="risk_band",
                    hover_name="district",
                    hover_data={"recent_3mo": ":.1f", "risk_score": ":.1f", "lat": False, "lng": False},
                    color_discrete_map={"Critical": "#ff5d6e", "High": "#ffb84d", "Moderate": "#a855f7", "Low": "#72e0b8"},
                    size_max=30,
                    zoom=6,
                    center={"lat": 15.3, "lon": 75.7},
                )
            fig.update_layout(map_style="carto-darkmatter", margin=dict(l=0, r=0, t=0, b=0), height=540, paper_bgcolor="#171126", legend=dict(bgcolor="rgba(0,0,0,0)"))
            st.plotly_chart(px_style(fig, height=540), width='stretch')

        with side_col:
            st.markdown(
                """
                <div class="activity-card">
                    <div class="eyebrow">Activities</div>
                    <div style="font-family:'Space Grotesk',sans-serif;font-size:22px;color:#f7f2ff;margin-top:4px;">Command activity</div>
                    <div class="activity-list">
                """,
                unsafe_allow_html=True,
            )
            for i, row in enumerate(top_categories.itertuples(index=False)):
                checked = i < 4
                st.markdown(
                    f"""
                    <div class="activity-item">
                        <div class="activity-label">{row.category}</div>
                        <div class="activity-check {'on' if checked else 'off'}">{'✓' if checked else '○'}</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            st.markdown("</div></div>", unsafe_allow_html=True)

        st.caption("Source: Karnataka crime register, Jan-2020 to Nov-2022. Geographic, network, and predictive layers elsewhere in this app are synthetic — see relevant tabs for detail.")


# ============================================================
# 2 · TAXONOMY EXPLORER
# ============================================================
elif section_num == 2:
    st.markdown('<div class="eyebrow">Situation Room / 02</div>', unsafe_allow_html=True)
    st.title("Crime Head Taxonomy Explorer")
    st.markdown('<div class="subtext">Search-as-you-scan replacement for the SCRB code book — every registered major head, section, and volume in one live table.</div>', unsafe_allow_html=True)
    st.write("")

    tax = b.taxonomy_table(300)
    search = st.text_input("Filter by crime head or section", "")
    if search:
        mask = (tax["Major Crime Head"] + " " + tax["Crime Head and Section"]).str.contains(search, case=False, na=False)
        tax = tax[mask]
    st.dataframe(tax.rename(columns={"Commits": "Recorded Commits"}), width='stretch', height=460, hide_index=True)

    st.subheader("Category Composition")
    top15 = b.top_categories(df, 15)
    fig = px.treemap(names=top15.index, parents=[""]*len(top15), values=top15.values,
                      color=top15.values, color_continuous_scale=[[0, "#162239"], [0.5, TEAL], [1, AMBER]])
    fig.update_layout(margin=dict(l=4, r=4, t=4, b=4))
    st.plotly_chart(px_style(fig, height=420), width='stretch')

# ============================================================
# 3 · GEOSPATIAL HOTSPOTS
# ============================================================
elif section_num == 3:
    st.markdown('<div class="eyebrow">Situation Room / 03</div>', unsafe_allow_html=True)
    st.title("Geospatial Hotspots")
    st.markdown('<div class="subtext">District-level drill-down with spatiotemporal clustering. <strong style="color:#ff8a3d">District, station, and coordinate detail are synthetic</strong> — the source file carries no location field.</div>', unsafe_allow_html=True)
    st.write("")

    fcol1, fcol2, fcol3 = st.columns([1, 1, 1])
    cats_avail = sorted(heat_df["category"].unique().tolist())
    months_avail = sorted(heat_df["month"].unique().tolist(), key=lambda m: pd.to_datetime(m, format="%b-%y"))
    cat_pick = fcol1.selectbox("Crime category", ["All categories"] + cats_avail)
    month_pick = fcol2.selectbox("Month", ["All months"] + months_avail)
    layer_pick = fcol3.radio("Layer", ["Density (heat)", "District markers"], horizontal=True)

    fdf = heat_df.copy()
    if cat_pick != "All categories":
        fdf = fdf[fdf["category"] == cat_pick]
    if month_pick != "All months":
        fdf = fdf[fdf["month"] == month_pick]

    map_col, detail_col = st.columns([1.5, 1])
    with map_col:
        if layer_pick.startswith("Density"):
            fig = px.density_map(
                fdf,
                lat="lat",
                lon="lng",
                z=None,
                radius=22,
                center={"lat": 15.3, "lon": 75.7},
                zoom=6,
                map_style="carto-darkmatter",
                color_continuous_scale=[[0, TEAL], [0.5, AMBER], [1, RED]],
            )
            fig.update_layout(margin=dict(l=0, r=0, t=0, b=0), height=540, paper_bgcolor="#111b2e")
        else:
            plot_df = ddf.copy()
            plot_df["color"] = np.where(plot_df["is_hotspot"], "Hotspot",
                                  np.where(plot_df["pct_change_vs_avg"] > 5, "Elevated", "Stable"))
            fig = px.scatter_map(plot_df, lat="lat", lon="lng", size="total", color="color",
                                     color_discrete_map={"Hotspot": RED, "Elevated": AMBER, "Stable": TEAL},
                                     hover_name="district",
                                     hover_data={"lat": False, "lng": False, "total": True, "pct_change_vs_avg": True},
                                     size_max=32, zoom=6, center={"lat": 15.3, "lon": 75.7})
            fig.update_layout(map_style="carto-darkmatter", margin=dict(l=0, r=0, t=0, b=0), height=540,
                               paper_bgcolor="#111b2e", legend=dict(bgcolor="rgba(0,0,0,0)"))
        st.plotly_chart(fig, width='stretch')
        st.caption("🔴 Hotspot (>25% vs trailing avg)  ·  🟠 Elevated  ·  🟢 Stable")

    with detail_col:
        st.subheader("District Drill-Down")
        pick_district = st.selectbox("Select district", ddf.sort_values("recent_3mo", ascending=False)["district"].tolist())
        drow = ddf[ddf["district"] == pick_district].iloc[0]
        if drow["is_hotspot"]:
            st.markdown(risk_badge("Critical") + " &nbsp;**HOTSPOT FLAGGED**", unsafe_allow_html=True)
        m1, m2 = st.columns(2)
        m1.metric("Total Recorded", f"{drow['total']:,.0f}")
        m2.metric("Recent 3-Month", f"{drow['recent_3mo']:,.0f}", f"{drow['pct_change_vs_avg']:+.1f}%")
        st.write(f"**Top category:** {drow['top_category']}")
        st.write(f"**Population (synthetic):** {drow['population_lakh']}L · {drow['urban_pct']}% urban")
        st.write(f"**Crime rate:** {drow['crime_rate_per_lakh']} / lakh population")
        st.markdown("**Station-level breakdown (synthetic)**")
        st_df = district_stations[pick_district]
        fig = px.bar(st_df, x="total", y="station", orientation="h", template=PLOTLY_TEMPLATE,
                     color_discrete_sequence=[TEAL])
        fig.update_layout(yaxis_title="", xaxis_title="")
        st.plotly_chart(px_style(fig, height=260), width='stretch')

    st.subheader("All Districts — Recent 3-Month Snapshot")
    show = ddf.sort_values("recent_3mo", ascending=False)[["district", "top_category", "recent_3mo", "pct_change_vs_avg", "is_hotspot"]].copy()
    show["status"] = np.where(show["is_hotspot"], "🔴 Hotspot", "🟢 Stable")
    st.dataframe(show.drop(columns="is_hotspot").rename(columns={
        "district": "District", "top_category": "Top Category", "recent_3mo": "Recent 3-mo",
        "pct_change_vs_avg": "vs Trailing Avg (%)", "status": "Status"}),
        width='stretch', height=340, hide_index=True)


# ============================================================
# 4 · HOTSPOT CLUSTERING (ML) — NEW FEATURE
# ============================================================
elif section_num == 4:
    st.markdown('<div class="eyebrow">Situation Room / 04</div>', unsafe_allow_html=True)
    st.title("Hotspot Clustering — K-Means")
    st.markdown('<div class="subtext">Unsupervised clustering groups districts by recent volume, momentum, and crime rate — an automated first pass at tiering districts for resource planning, instead of manually eyeballing a spreadsheet.</div>', unsafe_allow_html=True)
    st.write("")

    k = st.slider("Number of clusters", 2, 6, 4)
    clu = b.cluster_hotspots(k=k)

    fig = px.scatter(clu, x="pct_change_vs_avg", y="recent_3mo", color="cluster_label", size="crime_rate_per_lakh",
                      hover_name="district", template=PLOTLY_TEMPLATE,
                      color_discrete_sequence=PALETTE,
                      labels={"pct_change_vs_avg": "Momentum vs trailing avg (%)", "recent_3mo": "Recent 3-month volume"})
    st.plotly_chart(px_style(fig, height=440), width='stretch')

    st.subheader("Cluster Membership")
    for label in clu["cluster_label"].unique():
        sub = clu[clu["cluster_label"] == label].sort_values("risk_score", ascending=False)
        with st.expander(f"{label}  ·  {len(sub)} districts"):
            st.dataframe(sub[["district", "recent_3mo", "pct_change_vs_avg", "risk_score", "top_category"]]
                         .rename(columns={"district": "District", "recent_3mo": "Recent 3-mo",
                                           "pct_change_vs_avg": "Momentum %", "risk_score": "Risk Score",
                                           "top_category": "Top Category"}),
                         width='stretch', hide_index=True)
    st.caption("Clustering runs on synthetic district-level aggregates (see Geospatial tab for data provenance).")

# ============================================================
# 5 · TREND ANOMALIES (STATISTICAL)
# ============================================================
elif section_num == 5:
    st.markdown('<div class="eyebrow">Situation Room / 05</div>', unsafe_allow_html=True)
    st.title("Trend Anomalies — Statistical (Z-Score)")
    st.markdown('<div class="subtext">Flags months where a category breaks its own historical pattern (z-score &gt; 1.8) — the emerging-trend alert SCRB currently has no automated way to catch.</div>', unsafe_allow_html=True)
    st.write("")

    anomalies = b.zscore_anomalies(df)
    if len(anomalies) == 0:
        st.info("No statistically significant spikes detected in the current window.")
    else:
        for _, a in anomalies.iterrows():
            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:14px;padding:10px 14px;border:1px solid #243252;border-radius:5px;margin-bottom:8px;background:#111b2e;">
              <div style="width:8px;height:8px;border-radius:50%;background:#e15252;flex-shrink:0;"></div>
              <div style="flex:1;">
                <b>{a['category']}</b><br>
                <span class="confidential">{a['month']} · recorded {a['value']:.0f} vs historical avg {a['historical_avg']:.0f} · z={a['z_score']}</span>
              </div>
              <div style="font-family:'IBM Plex Mono',monospace;font-size:16px;color:#e15252;font-weight:600;">+{a['pct_above_avg']}%</div>
            </div>
            """, unsafe_allow_html=True)

    st.subheader("Category Trend Lines — top 6 by volume")
    top6 = b.top_categories(df, 6).index.tolist()
    labels, series = b.category_monthly(df, top6)
    cols = st.columns(3)
    for i, c in enumerate(top6):
        anomaly_months = set(anomalies[anomalies["category"] == c]["month"]) if len(anomalies) else set()
        vals = series[c]
        colors = [RED if m in anomaly_months else PALETTE[i % len(PALETTE)] for m in labels]
        sizes = [8 if m in anomaly_months else 3 for m in labels]
        fig = go.Figure(go.Scatter(x=labels, y=vals, mode="lines+markers", line=dict(color=PALETTE[i % len(PALETTE)], width=1.8),
                                    marker=dict(color=colors, size=sizes)))
        fig.update_layout(title=dict(text=c, font=dict(size=12)), showlegend=False, xaxis=dict(showticklabels=False))
        cols[i % 3].plotly_chart(px_style(fig, height=220), width='stretch')


# ============================================================
# 6 · ML ANOMALY DETECTION (Isolation Forest) — NEW FEATURE
# ============================================================
elif section_num == 6:
    st.markdown('<div class="eyebrow">Situation Room / 06</div>', unsafe_allow_html=True)
    st.title("Anomaly Detection — Isolation Forest (ML)")
    st.markdown('<div class="subtext">Where the z-score view checks one category at a time, this model looks at every top category <em>jointly</em> per month — catching unusual month-level combinations (e.g. several categories moving together) that single-series checks miss.</div>', unsafe_allow_html=True)
    st.write("")

    contamination = st.slider("Sensitivity (expected anomaly rate)", 0.03, 0.20, 0.08, 0.01)
    iso = b.isolation_forest_anomalies(df, contamination=contamination)

    fig = go.Figure()
    fig.add_trace(go.Bar(x=iso["month"], y=iso["anomaly_score"],
                          marker_color=[RED if f else "#2a3a5c" for f in iso["flag"]]))
    fig.update_layout(xaxis=dict(categoryorder="array", categoryarray=sorted(iso["month"], key=lambda m: pd.to_datetime(m, format="%b-%y"))),
                       yaxis_title="Anomaly score (higher = more unusual)")
    st.plotly_chart(px_style(fig, height=360), width='stretch')

    flagged = iso[iso["flag"]].sort_values("anomaly_score", ascending=False)
    st.subheader(f"Flagged Months ({len(flagged)})")
    st.dataframe(flagged.rename(columns={"month": "Month", "anomaly_score": "Anomaly Score", "flag": "Flagged"}),
                 width='stretch', hide_index=True)
    st.caption("Model: scikit-learn IsolationForest over the top-15 category monthly series, standardized jointly. Retrain cadence and contamination rate would be tuned against real investigator feedback in production.")


# ============================================================
# 7 · PREDICTIVE FORECASTING
# ============================================================
elif section_num == 7:
    st.markdown('<div class="eyebrow">Situation Room / 07</div>', unsafe_allow_html=True)
    st.title("Predictive Forecasting")
    st.markdown('<div class="subtext">Linear trend projection per category, 3 months out. Built to be swapped for a trained time-series model once enough labeled history exists — not a substitute for one.</div>', unsafe_allow_html=True)
    st.write("")

    top8 = b.top_categories(df, 8).index.tolist()
    cat_pick = st.selectbox("Category", top8)
    forecasts, future_labels = b.forecast_categories(df, top8)
    fc = forecasts[cat_pick]

    hist_x = fc["history_months"]
    fut_x = future_labels
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=hist_x, y=fc["history"], name="Recorded", line=dict(color=TEAL, width=2)))
    fig.add_trace(go.Scatter(x=[hist_x[-1]] + fut_x, y=[fc["history"][-1]] + fc["forecast"],
                              name="Projected", line=dict(color=AMBER, width=2, dash="dash")))
    st.plotly_chart(px_style(fig, height=400), width='stretch')

    c1, c2, c3 = st.columns(3)
    c1.metric("Trend slope (per month)", f"{fc['slope']:+.1f}")
    c2.metric(f"Projected {fut_x[0]}", f"{fc['forecast'][0]:.0f}")
    c3.metric(f"Projected {fut_x[-1]}", f"{fc['forecast'][-1]:.0f}")

    st.subheader("All Tracked Categories — Next 3 Months")
    rows = [{"Category": c, "Trend Slope": forecasts[c]["slope"],
             future_labels[0]: forecasts[c]["forecast"][0],
             future_labels[1]: forecasts[c]["forecast"][1],
             future_labels[2]: forecasts[c]["forecast"][2]} for c in top8]
    st.dataframe(pd.DataFrame(rows), width='stretch', hide_index=True)

# ============================================================
# 8 · NETWORK & LINK ANALYSIS
# ============================================================
elif section_num == 8:
    st.markdown('<div class="eyebrow">Situation Room / 08</div>', unsafe_allow_html=True)
    st.title("Network & Link Analysis")
    st.markdown('<div class="subtext">Node-based view connecting suspects, victims, and recurring locations. <strong style="color:#ff8a3d">Suspects/victims are synthetic</strong> (source file has no case-linkage data); MO tags are drawn from real registered minor-crime-head labels.</div>', unsafe_allow_html=True)
    st.write("")

    net = b.build_network()
    suspects = {n["id"]: n for n in net["nodes"] if n["type"] == "suspect"}
    crime_types = sorted({s["primary_crime"] for s in suspects.values()})

    fcol1, fcol2 = st.columns(2)
    crime_pick = fcol1.selectbox("Crime type", ["All"] + crime_types)
    mode_pick = fcol2.radio("Show", ["Full network", "Repeat offenders only"], horizontal=True)

    nodes = list(net["nodes"])
    edges = list(net["edges"])
    if crime_pick != "All":
        keep_s = {sid for sid, s in suspects.items() if s["primary_crime"] == crime_pick}
        edges = [e for e in edges if e["source"] in keep_s or e["target"] in keep_s]
        keep_ids = {e["source"] for e in edges} | {e["target"] for e in edges} | keep_s
        nodes = [n for n in nodes if n["id"] in keep_ids]
    if mode_pick == "Repeat offenders only":
        repeat_ids = {sid for sid, s in suspects.items() if s["repeat_offender"]}
        edges = [e for e in edges if e["source"] in repeat_ids or e["target"] in repeat_ids]
        keep_ids = {e["source"] for e in edges} | {e["target"] for e in edges}
        nodes = [n for n in nodes if n["id"] in keep_ids]

    G = nx.Graph()
    for n in nodes:
        G.add_node(n["id"], **n)
    for e in edges:
        if e["source"] in G and e["target"] in G:
            G.add_edge(e["source"], e["target"], **e)
    pos = nx.spring_layout(G, seed=42, k=0.6)

    edge_x, edge_y = [], []
    for u, v in G.edges():
        edge_x += [pos[u][0], pos[v][0], None]
        edge_y += [pos[u][1], pos[v][1], None]
    edge_trace = go.Scatter(x=edge_x, y=edge_y, mode="lines", line=dict(color="#243252", width=1), hoverinfo="none")

    type_color = {"suspect": AMBER, "victim": BLUE, "location": TEAL}
    node_x = [pos[n][0] for n in G.nodes()]
    node_y = [pos[n][1] for n in G.nodes()]
    node_color = [type_color[G.nodes[n]["type"]] for n in G.nodes()]
    node_size = [14 if G.nodes[n]["type"] == "suspect" and G.nodes[n].get("repeat_offender") else
                 10 if G.nodes[n]["type"] == "location" else 7 for n in G.nodes()]
    node_text = [G.nodes[n]["label"] for n in G.nodes()]
    node_trace = go.Scatter(x=node_x, y=node_y, mode="markers", text=node_text, hoverinfo="text",
                             marker=dict(color=node_color, size=node_size, line=dict(width=1, color="#0a111f")))

    fig = go.Figure([edge_trace, node_trace])
    fig.update_layout(showlegend=False, xaxis=dict(visible=False), yaxis=dict(visible=False))
    net_col, profile_col = st.columns([1.5, 1])
    with net_col:
        st.plotly_chart(px_style(fig, height=500), width='stretch')
        st.caption("🟠 Suspect · 🔵 Victim · 🟢 Location  ·  " +
                   f"{sum(1 for n in nodes if n['type']=='suspect')} suspects · "
                   f"{sum(1 for n in nodes if n['type']=='victim')} victims · {len(edges)} links")

    with profile_col:
        st.subheader("Suspect Profile")
        suspect_labels = {s["label"]: sid for sid, s in suspects.items() if sid in {n["id"] for n in nodes}}
        if suspect_labels:
            pick = st.selectbox("Select suspect", sorted(suspect_labels.keys()))
            s = suspects[suspect_labels[pick]]
            badge = risk_badge("High") if s["repeat_offender"] else risk_badge("Low")
            st.markdown((("**REPEAT OFFENDER** " + badge) if s["repeat_offender"] else ("Single incident " + badge)), unsafe_allow_html=True)
            st.write(f"**Suspect ID:** `{s['id']}`")
            st.write(f"**Home district:** {s['home_district']}")
            st.write(f"**Linked incidents:** {s['n_incidents']}")
            st.write(f"**Jurisdictions spanned:** {len(s['jurisdictions'])} — {', '.join(s['jurisdictions'])}")
            st.write(f"**Primary crime type:** {s['primary_crime']}")
            st.markdown("**MO tags:** " + " ".join(f'<span class="mo-tag">{t}</span>' for t in s["mo_tags"]), unsafe_allow_html=True)
            if len(s["jurisdictions"]) > 1:
                st.warning(f"⚠ Cross-jurisdiction pattern — same MO recurring across {len(s['jurisdictions'])} districts. Flag for inter-station correlation.")
        else:
            st.info("No suspects match the current filter.")


# ============================================================
# 9 · MO SIGNATURE MATCHING — NEW FEATURE
# ============================================================
elif section_num == 9:
    st.markdown('<div class="eyebrow">Situation Room / 09</div>', unsafe_allow_html=True)
    st.title("MO Signature Matching")
    st.markdown('<div class="subtext">Finds suspects who share 2+ identical modus-operandi tags despite operating out of <em>different</em> home districts — a lightweight stand-in for the cross-jurisdiction pattern matching investigators currently do by memory and phone calls between stations.</div>', unsafe_allow_html=True)
    st.write("")

    matches = b.mo_signature_matches()
    if len(matches) == 0:
        st.info("No cross-district MO matches found in the current synthetic network.")
    else:
        st.metric("Cross-District MO Matches Found", len(matches))
        for _, m in matches.iterrows():
            st.markdown(f"""
            <div style="padding:12px 16px;border:1px solid #243252;border-radius:5px;margin-bottom:8px;background:#111b2e;">
              <div style="display:flex;justify-content:space-between;">
                <b>{m['suspect_a']}</b> ({m['district_a']}) &nbsp;↔&nbsp; <b>{m['suspect_b']}</b> ({m['district_b']})
                <span style="font-family:'IBM Plex Mono',monospace;color:#ff9f7a;">{m['n_shared']} shared MO tags</span>
              </div>
              <div class="confidential" style="margin-top:6px;">Shared: {m['shared_mo']}</div>
            </div>
            """, unsafe_allow_html=True)
    st.caption("Matching runs on the synthetic suspect network's MO tags, which are themselves drawn from real registered minor-crime-head labels.")


# ============================================================
# 10 · SOCIO-ECONOMIC OVERLAY
# ============================================================
elif section_num == 10:
    st.markdown('<div class="eyebrow">Situation Room / 10</div>', unsafe_allow_html=True)
    st.title("Socio-Economic Correlation Overlay")
    st.markdown('<div class="subtext">Overlays crime rate against urbanization and literacy to explore the "why" behind the "where". <strong style="color:#ff8a3d">Urbanization, literacy, and population figures are synthetic</strong> illustrative indices, not official Census/NCRB statistics.</div>', unsafe_allow_html=True)
    st.write("")

    socio = b.socioeconomic_table()
    x_axis = st.selectbox("X-axis indicator", ["urban_pct", "literacy_pct", "population_lakh"],
                           format_func=lambda x: {"urban_pct": "Urbanization %", "literacy_pct": "Literacy %", "population_lakh": "Population (lakh)"}[x])

    try:
        fig = px.scatter(socio, x=x_axis, y="crime_rate_per_lakh", color="risk_band", size="population_lakh",
                         hover_name="district", template=PLOTLY_TEMPLATE,
                         color_discrete_map={"Critical": RED, "High": AMBER, "Moderate": BLUE, "Low": TEAL},
                         trendline="ols")
    except Exception:
        fig = px.scatter(socio, x=x_axis, y="crime_rate_per_lakh", color="risk_band", size="population_lakh",
                         hover_name="district", template=PLOTLY_TEMPLATE,
                         color_discrete_map={"Critical": RED, "High": AMBER, "Moderate": BLUE, "Low": TEAL})
        st.caption("Trendline disabled because the optional statsmodels dependency is not available in this environment.")

    st.plotly_chart(px_style(fig, height=440), width='stretch')

    corr = socio[["urban_pct", "literacy_pct", "population_lakh", "crime_rate_per_lakh"]].corr()["crime_rate_per_lakh"].drop("crime_rate_per_lakh")
    st.subheader("Correlation with Crime Rate")
    cc1, cc2, cc3 = st.columns(3)
    cc1.metric("Urbanization %", f"{corr['urban_pct']:+.2f}")
    cc2.metric("Literacy %", f"{corr['literacy_pct']:+.2f}")
    cc3.metric("Population", f"{corr['population_lakh']:+.2f}")
    st.caption("Pearson correlation, synthetic indicators. In production this layer would pull real Census/NCRB socio-economic data joined at the district level.")

    st.dataframe(socio.rename(columns={"district": "District", "urban_pct": "Urban %", "literacy_pct": "Literacy %",
                                        "population_lakh": "Population (L)", "crime_rate_per_lakh": "Crime Rate/Lakh",
                                        "risk_band": "Risk Band"}), width='stretch', hide_index=True)


# ============================================================
# 11 · RESOURCE DEPLOYMENT RECOMMENDER — NEW FEATURE
# ============================================================
elif section_num == 11:
    st.markdown('<div class="eyebrow">Situation Room / 11</div>', unsafe_allow_html=True)
    st.title("Resource Deployment Recommender")
    st.markdown('<div class="subtext">Turns the risk score into an actual allocation: distributes a fixed pool of patrol/investigation units across districts, weighted by risk, with a floor so no district goes uncovered. This is the "proactive policing" output the brief calls for — a starting point for duty officers, not a final order.</div>', unsafe_allow_html=True)
    st.write("")

    n_units = st.slider("Total units available this cycle", 10, 60, 30)
    dep = b.deployment_recommendations(n_units)

    fig = px.bar(dep.sort_values("recommended_units", ascending=True), x="recommended_units", y="district",
                 orientation="h", color="risk_band", template=PLOTLY_TEMPLATE,
                 color_discrete_map={"Critical": RED, "High": AMBER, "Moderate": BLUE, "Low": TEAL})
    fig.update_layout(yaxis_title="", xaxis_title="Recommended units")
    st.plotly_chart(px_style(fig, height=700), width='stretch')

    st.subheader("Deployment Table")
    show = dep.copy()
    show["Risk Band"] = show["risk_band"].apply(lambda x: x)
    st.dataframe(show.rename(columns={"district": "District", "risk_score": "Risk Score",
                                       "recommended_units": "Recommended Units", "shift_focus": "Shift Focus",
                                       "top_category": "Top Category"}).drop(columns=["risk_band"]),
                 width='stretch', hide_index=True)
    st.caption(f"Total allocated: {dep['recommended_units'].sum()} of {n_units} units. Allocation = proportional to composite risk score, floor of 1 unit/district. Swap in real duty-roster constraints for production use.")

# ============================================================
# 12 · OPERATIONS COMMAND CENTER
# ============================================================
elif section_num == 12:
    st.markdown('<div class="eyebrow">Operations Control / 12</div>', unsafe_allow_html=True)
    st.title("Operations Command Center")
    st.markdown('<div class="subtext">A realistic control-room view that combines incident readiness, live intelligence, and deployment guidance for supervisors and station heads.</div>', unsafe_allow_html=True)
    st.write("")

    st.markdown("""
    <div class="hero-panel">
      <div class="eyebrow">Shift brief</div>
      <h3 style="margin: 4px 0 6px;">Priority focus for the next 6 hours</h3>
      <div class="subtext">Review district stress levels, assign patrol coverage, and escalate emerging anomalies before they become serious incidents.</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown(
        f"""
        <div class="info-card">
            <div class="eyebrow">Personnel dashboard</div>
            <div style="font-size: 18px; color: #f5f7fb; font-weight: 700; margin-top: 4px;">{st.session_state.current_user.upper() if st.session_state.current_user else 'OFFICER'}</div>
            <div class="subtext" style="margin-top: 4px;">{st.session_state.user_role or 'Role N/A'} · {st.session_state.unit or 'Unit N/A'}</div>
            <div style="margin-top: 10px; display: flex; flex-wrap: wrap; gap: 8px;">
                <span class="user-pill">Grade: {st.session_state.police_grade or 'N/A'}</span>
                <span class="user-pill">Level: {st.session_state.police_level or 'N/A'}</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("""
        <div class="info-card">
          <div class="eyebrow">Incident queue</div>
          <div style="font-size: 28px; color: #ff8a3d; font-weight: 700;">12</div>
          <div class="subtext">Open cases awaiting supervisory review.</div>
        </div>
        """, unsafe_allow_html=True)
    with c2:
        st.markdown("""
        <div class="info-card">
          <div class="eyebrow">High-priority districts</div>
          <div style="font-size: 28px; color: #e15252; font-weight: 700;">4</div>
          <div class="subtext">Districts currently above the alert threshold.</div>
        </div>
        """, unsafe_allow_html=True)
    with c3:
        st.markdown("""
        <div class="info-card">
          <div class="eyebrow">Patrol readiness</div>
          <div style="font-size: 28px; color: #3fa9a0; font-weight: 700;">82%</div>
          <div class="subtext">Coverage available for the next deployment cycle.</div>
        </div>
        """, unsafe_allow_html=True)

    st.subheader("Suggested deployment actions")
    dep = b.deployment_recommendations(18)
    st.dataframe(dep.head(8).rename(columns={"district": "District", "risk_score": "Risk Score", "risk_band": "Risk Band",
                                             "recommended_units": "Recommended Units", "shift_focus": "Shift Focus",
                                             "top_category": "Top Category"}), width='stretch', hide_index=True)

    st.subheader("Operational notes")
    st.info("Use the anomaly, hotspot, and forecasting views together to brief the duty officer, then move to resource deployment for a recommended allocation plan.")


# ============================================================
# 13 · DUPLICATE COMPLAINT DETECTION
# ============================================================
elif section_num == 13:
    st.markdown('<div class="eyebrow">Data Quality / 13</div>', unsafe_allow_html=True)
    st.title("Duplicate Complaint Detection")
    st.markdown('<div class="subtext">Checks repeated complaint numbers, exact duplicate rows, missing values, and malformed records so the intake queue can be cleaned before analysis.</div>', unsafe_allow_html=True)
    st.write("")

    audit = build_quality_audit(df)
    top_col1, top_col2, top_col3, top_col4 = st.columns(4)
    top_col1.metric("Duplicate Complaint Numbers", audit["duplicate_complaint_rows"])
    top_col2.metric("Repeated Records", audit["repeated_record_rows"])
    top_col3.metric("Missing Values", audit["missing_value_total"])
    top_col4.metric("Invalid Data Rows", audit["invalid_data_rows"])

    left_col, right_col = st.columns([1.15, 0.85])
    with left_col:
        st.subheader("Duplicate complaint numbers")
        if len(audit["duplicate_counts"]):
            st.dataframe(audit["duplicate_counts"], width='stretch', height=320, hide_index=True)
        else:
            st.success("No duplicate complaint numbers found.")

        st.subheader("Repeated records")
        if len(audit["repeated_records"]):
            st.dataframe(audit["repeated_records"].head(50), width='stretch', height=280, hide_index=True)
        else:
            st.success("No repeated rows found.")

    with right_col:
        st.subheader("Missing values by column")
        fig = px.bar(audit["missing_by_column"], x="Column", y="Missing Values", color="Missing Values",
                     color_continuous_scale=[[0, TEAL], [1, RED]], template=PLOTLY_TEMPLATE)
        fig.update_layout(xaxis_title="", yaxis_title="Missing values")
        st.plotly_chart(px_style(fig, height=360), width='stretch')

        st.subheader("Invalid rows")
        if len(audit["invalid_rows"]):
            st.dataframe(audit["invalid_rows"][ ["Complaint Number", "Major Crime Head", "Crime Head and Section", "Minor Crime Head", "Commits", "Month", "Issue"] ], width='stretch', height=260, hide_index=True)
        else:
            st.success("No invalid rows found.")


# ============================================================
# 14 · CRIME REPORT GENERATOR
# ============================================================
# 14 · CRIME REPORT GENERATOR
elif section_num == 14:
    st.markdown('<div class="eyebrow">Reporting / 14</div>', unsafe_allow_html=True)
    st.title("Crime Report Generator")
    st.markdown('<div class="subtext">Creates downloadable PDF, Excel, and CSV reports with charts, KPI highlights, and data-quality insights from the current dataset.</div>', unsafe_allow_html=True)
    st.write("")

    bundle = build_report_bundle(df, meta)
    pdf_bytes = build_pdf_report(bundle)
    xlsx_bytes = build_excel_report(bundle)
    csv_bytes = build_csv_summary(bundle)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Summary Rows", len(bundle["summary_rows"]))
    c2.metric("Insight Cards", len(bundle["kpi"]["insights"]))
    c3.metric("Quality Checks", len(bundle["quality"]["issue_summary"]))
    c4.metric("Crime Categories", bundle["kpi"]["crime_categories"])

    st.subheader("Key insights")
    for insight in bundle["kpi"]["insights"]:
        st.markdown(f"- {insight}")

    left_col, right_col = st.columns(2)
    with left_col:
        st.subheader("Monthly crime trend")
        fig = px.line(bundle["kpi"]["monthly"], x="month", y="total", markers=True, template=PLOTLY_TEMPLATE)
        fig.update_traces(line_color=TEAL)
        fig.update_layout(xaxis_title="Month", yaxis_title="Total crimes")
        st.plotly_chart(px_style(fig, height=330), width='stretch')
    with right_col:
        st.subheader("Top crime categories")
        fig = px.bar(bundle["kpi"]["top_categories"], x="Total Crimes", y="Crime Category", orientation="h",
                     template=PLOTLY_TEMPLATE, color="Total Crimes", color_continuous_scale=[[0, BLUE], [1, AMBER]])
        fig.update_layout(yaxis_title="", xaxis_title="Total crimes")
        st.plotly_chart(px_style(fig, height=330), width='stretch')

    st.subheader("Download report files")
    d1, d2, d3 = st.columns(3)
    d1.download_button("Download PDF report", pdf_bytes, file_name="scrb_crime_report.pdf", mime="application/pdf", use_container_width=True)
    d2.download_button("Download Excel report", xlsx_bytes, file_name="scrb_crime_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    d3.download_button("Download CSV summary", csv_bytes, file_name="scrb_crime_summary.csv", mime="text/csv", use_container_width=True)


# ============================================================
# 15 · KPI DASHBOARD
# ============================================================
elif section_num == 15:
    st.markdown('<div class="eyebrow">Performance / 15</div>', unsafe_allow_html=True)
    st.title("KPI Dashboard")
    st.markdown('<div class="subtext">Tracks the headline indicators leadership asked for: complaints, crime volume, peak month, fastest-growing crime, average monthly crimes, and category breadth.</div>', unsafe_allow_html=True)
    st.write("")

    kpi = build_kpi_snapshot(df)
    c1, c2, c3 = st.columns(3)
    c1.metric("Total Complaints", f"{kpi['total_complaints']:,}")
    c2.metric("Total Crimes", f"{kpi['total_crimes']:.0f}")
    c3.metric("Number of Crime Categories", kpi["crime_categories"])
    c4, c5 = st.columns(2)
    c4.metric("Highest Crime Month", f"{kpi['highest_month']['month']}", f"{int(kpi['highest_month']['total'])} crimes")
    c5.metric("Average Monthly Crimes", f"{kpi['avg_monthly']:.1f}")

    g1, g2 = st.columns(2)
    with g1:
        st.subheader("Monthly crime trend")
        fig = px.bar(kpi["monthly"], x="month", y="total", template=PLOTLY_TEMPLATE, color="total", color_continuous_scale=[[0, TEAL], [1, AMBER]])
        fig.update_layout(xaxis_title="Month", yaxis_title="Crimes")
        st.plotly_chart(px_style(fig, height=340), width='stretch')
    with g2:
        st.subheader("Fastest growing crimes")
        if len(kpi["growth_table"]):
            fig = px.bar(kpi["growth_table"].head(8), x="Growth %", y="Crime Category", orientation="h", template=PLOTLY_TEMPLATE,
                         color="Growth %", color_continuous_scale=[[0, BLUE], [1, RED]])
            fig.update_layout(yaxis_title="", xaxis_title="Growth %")
            st.plotly_chart(px_style(fig, height=340), width='stretch')
            st.dataframe(kpi["growth_table"], width='stretch', hide_index=True)
        else:
            st.info("Not enough history to compute growth rates.")


# ============================================================
# 16 · DATA QUALITY DASHBOARD
# ============================================================
elif section_num == 16:
    st.markdown('<div class="eyebrow">Data Governance / 16</div>', unsafe_allow_html=True)
    st.title("Data Quality Dashboard")
    st.markdown('<div class="subtext">Monitors missing values, duplicate entries, invalid months, empty crime categories, and section inconsistencies so the source file can be cleaned before analysis. Use the issue desk below to file a report directly from this view.</div>', unsafe_allow_html=True)
    st.write("")

    audit = build_quality_audit(df)
    q1, q2, q3, q4 = st.columns(4)
    q1.metric("Missing Values", audit["missing_value_total"])
    q2.metric("Duplicate Entries", audit["repeated_record_rows"])
    q3.metric("Invalid Months", audit["invalid_month_rows"])
    q4.metric("Section Inconsistencies", audit["section_inconsistency_rows"])

    left_col, right_col = st.columns([1, 1])
    with left_col:
        st.subheader("Issue summary")
        st.dataframe(audit["issue_summary"], width='stretch', hide_index=True)

        st.subheader("Missing values by column")
        fig = px.bar(audit["missing_by_column"], x="Column", y="Missing Values", template=PLOTLY_TEMPLATE, color="Missing Values",
                     color_continuous_scale=[[0, TEAL], [1, RED]])
        fig.update_layout(xaxis_title="", yaxis_title="Missing values")
        st.plotly_chart(px_style(fig, height=320), width='stretch')

    with right_col:
        st.subheader("Duplicate complaint numbers")
        if len(audit["duplicate_counts"]):
            st.dataframe(audit["duplicate_counts"], width='stretch', height=180, hide_index=True)
        else:
            st.success("No duplicate complaint numbers found.")

        st.subheader("Invalid rows")
        if len(audit["invalid_rows"]):
            st.dataframe(audit["invalid_rows"][["Complaint Number", "Major Crime Head", "Crime Head and Section", "Minor Crime Head", "Commits", "Month", "Issue"]], width='stretch', height=240, hide_index=True)
        else:
            st.success("No invalid rows found.")

    st.subheader("Quality report and issue desk")
    report_col, issue_col = st.columns([0.95, 1.05])
    with report_col:
        st.markdown("<div class='info-card'>", unsafe_allow_html=True)
        st.markdown("<div class='eyebrow'>Take report</div>", unsafe_allow_html=True)
        st.markdown("<div class='subtext'>Download a compact quality report for review, escalation, or handoff.</div>", unsafe_allow_html=True)
        st.download_button(
            "Download quality report",
            build_quality_report_text(audit),
            file_name="scrb_data_quality_report.txt",
            mime="text/plain",
            use_container_width=True,
        )
        st.download_button(
            "Download issue summary CSV",
            audit["issue_summary"].to_csv(index=False).encode("utf-8"),
            file_name="scrb_data_quality_issues.csv",
            mime="text/csv",
            use_container_width=True,
        )
        st.markdown("</div>", unsafe_allow_html=True)

    with issue_col:
        st.markdown("<div class='info-card'>", unsafe_allow_html=True)
        st.markdown("<div class='eyebrow'>File an issue</div>", unsafe_allow_html=True)
        with st.form("data_quality_issue_form"):
            issue_type = st.selectbox(
                "Issue type",
                ["Duplicate complaint number", "Repeated record", "Missing values", "Invalid month", "Empty crime category", "Section inconsistency", "Invalid data"],
            )
            priority = st.selectbox("Priority", ["Low", "Medium", "High", "Critical"])
            summary = st.text_input("Short summary", placeholder="Describe the problem in one line")
            details = st.text_area("Details", placeholder="Add context, row references, or next steps")
            affected_rows = st.number_input("Affected rows", min_value=0, value=1, step=1)
            submitted = st.form_submit_button("Create issue ticket", use_container_width=True)

        if submitted:
            ticket_id, ticket_df = build_issue_ticket(issue_type, priority, summary, details, int(affected_rows))
            if "filed_issues" not in st.session_state:
                st.session_state.filed_issues = []
            st.session_state.filed_issues.append(ticket_df.iloc[0].to_dict())
            st.success(f"Issue ticket created: {ticket_id}")
            st.download_button(
                "Download issue ticket CSV",
                ticket_df.to_csv(index=False).encode("utf-8"),
                file_name=f"{ticket_id.lower()}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        if st.session_state.get("filed_issues"):
            st.markdown("**Filed issues**")
            st.dataframe(pd.DataFrame(st.session_state.filed_issues), width='stretch', hide_index=True)
        st.markdown("</div>", unsafe_allow_html=True)


# ============================================================
# 17 · DATASET AGENT ANALYSIS
# ============================================================
elif section_num == 17:
    st.markdown('<div class="eyebrow">Agent Desk / 17</div>', unsafe_allow_html=True)
    st.title("Dataset Agent Analysis")
    st.markdown('<div class="subtext">Choose a specific slice of the dataset, or upload a new CSV and analyze that file directly like an analyst. Use this when you want focused answers about one crime head, section, minor head, or month.</div>', unsafe_allow_html=True)
    st.write("")

    if "uploaded_dataset_df" not in st.session_state:
        st.session_state.uploaded_dataset_df = None
    if "uploaded_dataset_name" not in st.session_state:
        st.session_state.uploaded_dataset_name = "Built-in Crime_Data.csv"
    if "agent_analysis" not in st.session_state:
        st.session_state.agent_analysis = b.analyze_dataset_slice(df, query="overall dataset overview")

    uploaded_file = st.file_uploader("Upload a CSV dataset", type=["csv"], help="The file should include Complaint Number, Major Crime Head, Crime Head and Section, Minor Crime Head, Commits, and Month.")
    if uploaded_file is not None:
        try:
            uploaded_df = pd.read_csv(uploaded_file)
            uploaded_df = b.prepare_dataset_frame(uploaded_df)
            st.session_state.uploaded_dataset_df = uploaded_df
            st.session_state.uploaded_dataset_name = uploaded_file.name
            st.success(f"Loaded dataset: {uploaded_file.name} ({len(uploaded_df):,} rows)")
        except Exception as exc:
            st.session_state.uploaded_dataset_df = None
            st.error(f"Could not load the uploaded dataset: {exc}")

    active_df = st.session_state.uploaded_dataset_df if st.session_state.uploaded_dataset_df is not None else df
    st.markdown(
        f"""
        <div class="info-card" style="margin-top: 10px;">
          <div class="confidential">ACTIVE DATASET</div>
          <div style="font-size: 14px; color: #f5f7fb; margin-top: 4px;">{st.session_state.uploaded_dataset_name}</div>
          <div style="margin-top: 6px; color: #b5c3dd; font-size: 13px;">Rows available: {len(active_df):,}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    major_options = ["All Major Crime Heads"] + sorted(active_df["Major Crime Head"].dropna().astype(str).unique().tolist())
    section_options = ["All Sections"] + sorted(active_df["Crime Head and Section"].dropna().astype(str).unique().tolist())
    minor_options = ["All Minor Crime Heads"] + sorted(active_df["Minor Crime Head"].dropna().astype(str).unique().tolist())
    month_options = ["All Months"] + sorted(
        active_df["Month"].dropna().astype(str).unique().tolist(),
        key=lambda m: pd.to_datetime(m, format="%b-%y")
    )

    with st.form("agent_analysis_form"):
        f1, f2, f3, f4 = st.columns(4)
        major_pick = f1.selectbox("Major crime head", major_options)
        section_pick = f2.selectbox("Crime section", section_options)
        minor_pick = f3.selectbox("Minor crime head", minor_options)
        month_pick = f4.selectbox("Month", month_options)
        question = st.text_area(
            "Ask the agent",
            placeholder="Example: Why is this slice rising? Check for duplicates and section concentration.",
        )
        run_analysis = st.form_submit_button("Analyze slice", use_container_width=True)

    if run_analysis:
        st.session_state.agent_analysis = b.analyze_dataset_slice(
            active_df,
            major=major_pick,
            section=section_pick,
            minor=minor_pick,
            month=month_pick,
            query=question,
        )

    quick_actions = st.columns(2)
    with quick_actions[0]:
        if st.button("Analyze full active dataset", use_container_width=True):
            st.session_state.agent_analysis = b.analyze_dataset_slice(active_df, query=question or "full dataset overview")
            st.rerun()
    with quick_actions[1]:
        if st.button("Reset to built-in dataset", use_container_width=True):
            st.session_state.uploaded_dataset_df = None
            st.session_state.uploaded_dataset_name = "Built-in Crime_Data.csv"
            st.session_state.agent_analysis = b.analyze_dataset_slice(df, query="overall dataset overview")
            st.rerun()

    result = st.session_state.agent_analysis
    summary = result["summary"]

    s1, s2, s3, s4 = st.columns(4)
    s1.metric("Rows in slice", f"{summary['rows']:,}")
    s2.metric("Commits", f"{summary['total_commits']:.1f}")
    s3.metric("Peak month", summary["peak_month"], f"{summary['peak_month_total']:.1f}")
    s4.metric("Trend", summary["trend_label"], f"{summary['recent_change_pct']:+.1f}%")

    response_col, stats_col = st.columns([1.2, 0.8])
    with response_col:
        st.markdown(
            f"""
            <div class="info-card">
              <div class="eyebrow">Agent response</div>
              <div style="font-size: 15px; line-height: 1.6; color: #f5f7fb;">{result['response']}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if summary["duplicate_rows"]:
            st.warning(f"This slice includes {summary['duplicate_rows']} rows tied to repeated complaint numbers.")

        st.subheader("Key findings")
        for item in result["key_findings"]:
            st.markdown(f"- {item}")

        st.subheader("Filtered records preview")
        preview_cols = [col for col in ["Complaint Number", "Major Crime Head", "Crime Head and Section", "Minor Crime Head", "Commits", "Month"] if col in result["slice_df"].columns]
        st.dataframe(result["slice_df"][preview_cols].head(120), width='stretch', height=320, hide_index=True)

        st.download_button(
            "Download filtered slice CSV",
            result["slice_df"].to_csv(index=False).encode("utf-8"),
            file_name="scrb_agent_slice.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with stats_col:
        st.subheader("Monthly trend")
        if len(result["monthly_trend"]):
            fig = px.line(result["monthly_trend"], x="month", y="total", markers=True, template=PLOTLY_TEMPLATE)
            fig.update_traces(line_color=RED, marker=dict(size=6))
            fig.update_layout(xaxis_title="Month", yaxis_title="Commits")
            st.plotly_chart(px_style(fig, height=280), width='stretch')
        else:
            st.info("No monthly trend available for this slice.")

        st.subheader("Top major heads")
        if len(result["top_major"]):
            fig = px.bar(result["top_major"], x="Total Commits", y="Major Crime Head", orientation="h", template=PLOTLY_TEMPLATE,
                         color="Total Commits", color_continuous_scale=[[0, "#2a1320"], [1, RED]])
            fig.update_layout(xaxis_title="Commits", yaxis_title="")
            st.plotly_chart(px_style(fig, height=280), width='stretch')

    lower_left, lower_right = st.columns(2)
    with lower_left:
        st.subheader("Top minor heads")
        if len(result["top_minor"]):
            fig = px.bar(result["top_minor"], x="Total Commits", y="Minor Crime Head", orientation="h", template=PLOTLY_TEMPLATE,
                         color="Total Commits", color_continuous_scale=[[0, "#2a1320"], [1, AMBER]])
            fig.update_layout(xaxis_title="Commits", yaxis_title="")
            st.plotly_chart(px_style(fig, height=320), width='stretch')

    with lower_right:
        st.subheader("Top sections")
        if len(result["top_sections"]):
            fig = px.bar(result["top_sections"], x="Total Commits", y="Crime Head and Section", orientation="h", template=PLOTLY_TEMPLATE,
                         color="Total Commits", color_continuous_scale=[[0, "#2a1320"], [1, VIOLET]])
            fig.update_layout(xaxis_title="Commits", yaxis_title="")
            st.plotly_chart(px_style(fig, height=320), width='stretch')
